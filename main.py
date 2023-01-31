from operator import index
import openpyxl
from numpy.lib.npyio import load
from openpyxl import load_workbook
import win32com.client
import pandas as pd
from pandas.core.interchange import column
import tkinter as tk
import tkinter.filedialog
from tkinter import messagebox
import os
import sys

#보고서 파일을 불러들인다.
messagebox.showwarning("보고서", "보고서 파일을 선택하세요.")    #파일 선택 안했을 때 메세지 출력

report_path =tk.filedialog.askopenfilename(initialdir = "os.path.expanduser('~')",
                                        title = "보고서 파일을 선택하세요.", filetypes = (("xlsx","*.xlsx"),("all files","*.*")))
#files 변수에 선택 파일 경로 넣기
if report_path == '':
        messagebox.showwarning("경고", "파일을 추가 하세요")    #파일 선택 안했을 때 메세지 출력
        sys.exit('종료')


#파일철_엑셀 파일을 불러들인다.
messagebox.showwarning("파일철_엑셀", "파일철_엑셀을 선택하세요.")    #파일 선택 안했을 때 메세지 출력
sheet_lable_path =tk.filedialog.askopenfilename(initialdir = "os.path.expanduser('~')",
                                        title = "파일철_엑셀을 선택하세요.", filetypes = (("xlsx","*.xlsx"),("all files","*.*")))
#files 변수에 선택 파일 경로 넣기
if sheet_lable_path == '':
        messagebox.showwarning("경고", "파일을 추가 하세요")    #파일 선택 안했을 때 메세지 출력
        sys.exit('종료')
#print(filename) #선택된 파일의 경로 및 파일명
#sheet_lable_path =os.path.dirname(filename) #파일경로에서, 폴더 경로만 가지고 오기

#
#1. 행 개수 가져옴
wb = load_workbook(report_path)
ws = wb['결과보고서']

max_row = ws.max_row
#print(max_row)
##################################
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False

wb = excel.Workbooks.Open(report_path)
ws = wb.Worksheets("결과보고서")

ws.Range("G18:G"+str(max_row)).Copy()
ws.Range("G18:G"+str(max_row)).PasteSpecial(12)
#
#
ws.Range("H18:H"+str(max_row)).Copy()
ws.Range("H18:H"+str(max_row)).PasteSpecial(12)
#
#
ws.Range("I18:I"+str(max_row)).Copy()
ws.Range("I18:I"+str(max_row)).PasteSpecial(12)
#
temp_result = r"c:\temp_result.xlsx"
ws.SaveAs(temp_result)
excel.Quit()
#
df = pd.read_excel(temp_result, skiprows=16,usecols=[1,6,7,8])

miss_tag = pd.read_excel(temp_result, skiprows=16,usecols=[1,6], index_col=False)
miss_tag.columns = ['운영부서','수량']
miss_tag = miss_tag.loc[miss_tag['수량'] != 0]
miss_tag = miss_tag.reset_index(drop=True)

binder_tag = pd.read_excel(temp_result, skiprows=16,usecols=[1,7], index_col=False)
binder_tag.columns = ['운영부서','수량']
binder_tag = binder_tag.loc[binder_tag['수량'] != 0]
binder_tag = binder_tag.reset_index(drop=True)

disposal_tag = pd.read_excel(temp_result, skiprows=16,usecols=[1,8], index_col=False)
disposal_tag.columns = ['운영부서','수량']
disposal_tag = disposal_tag.loc[disposal_tag['수량'] != 0]
disposal_tag = disposal_tag.reset_index(drop=True)

wb = openpyxl.load_workbook(sheet_lable_path)

ws = wb['미부착']
ws2 = wb['파일철']
ws3 = wb['폐기']

temp = 0
for i in range (0,len(miss_tag['운영부서']),1):
    ws.cell(row = i + 5, column = 2).value = miss_tag.loc[i,"운영부서"]
    ws.cell(row = i + 5, column = 3).value = miss_tag.loc[i, "수량"]
    temp = temp + miss_tag.loc[i,"수량"]
    #print(miss_tag.loc[i,"운영부서"])
ws["B2"] = "("+str(temp)+"점)"
#
temp = 0
for i in range (0,len(binder_tag['운영부서']),1):
    ws2.cell(row = i + 5, column = 2).value = binder_tag.loc[i,"운영부서"]
    ws2.cell(row = i + 5, column = 3).value = binder_tag.loc[i, "수량"]
    temp = temp + binder_tag.loc[i,"수량"]
    #print(miss_tag.loc[i,"운영부서"])
ws2["B2"] = "("+str(temp)+"점)"
#
temp = 0
for i in range (0,len(disposal_tag['운영부서']),1):
    ws3.cell(row = i + 5, column = 2).value = disposal_tag.loc[i,"운영부서"]
    ws3.cell(row = i + 5, column = 3).value = disposal_tag.loc[i, "수량"]
    temp = temp + disposal_tag.loc[i,"수량"]
    #print(miss_tag.loc[i,"운영부서"])
ws3["B2"] = "("+str(temp)+"점)"

wb.save("파일철_엑셀_완성.xlsx")
os.remove(temp_result)